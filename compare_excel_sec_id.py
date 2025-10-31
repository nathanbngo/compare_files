import argparse
import os
import sys
from typing import Dict, List, Set, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import numpy as np
import time


RED_FILL = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type="solid")
BLUE_FILL = PatternFill(start_color="FF99CCFF", end_color="FF99CCFF", fill_type="solid")


def read_excel_as_str(path: str) -> pd.DataFrame:
    print(f"Reading Excel: {path}")
    df = pd.read_excel(path, dtype=str)
    # Normalize whitespace and NaNs to empty strings for stable comparison
    df = df.fillna("")
    for col in df.columns:
        df[col] = df[col].astype(str).str.strip()
    print(f"Loaded {len(df)} rows, {len(df.columns)} columns from: {path}")
    return df


def get_sec_id_column_name(df: pd.DataFrame) -> str:
    if len(df.columns) == 0:
        raise ValueError("Excel has no columns")
    # Prefer an explicit 'SEC_ID' match (case-insensitive), else use first column
    for col in df.columns:
        if str(col).strip().lower() == "sec_id":
            return col
    return df.columns[0]


def build_row_maps(df: pd.DataFrame, id_col: str) -> Dict[str, int]:
    id_to_index: Dict[str, int] = {}
    for idx, val in df[id_col].items():
        key = str(val)
        if key not in id_to_index:
            id_to_index[key] = idx
    return id_to_index


def intersect_columns(df1: pd.DataFrame, df2: pd.DataFrame, exclude: Set[str]) -> List[str]:
    cols = [c for c in df1.columns if c in df2.columns and c not in exclude]
    return cols


def compare_rows(
    df1: pd.DataFrame,
    df2: pd.DataFrame,
    id_col1: str,
    id_col2: str,
) -> Tuple[Dict[str, List[str]], Dict[str, List[str]], Set[str], Set[str]]:
    print("Comparing rows by SEC_ID...")
    start_ts = time.perf_counter()

    # Fast path: two-pointer merge walk if both SEC_ID columns are ascending
    def _is_sorted(series: pd.Series) -> bool:
        vals = series.to_numpy()
        # Compare adjacent values as strings (already normalized to str)
        return np.all(vals[1:] >= vals[:-1]) if len(vals) > 1 else True

    use_fast_path = _is_sorted(df1[id_col1]) and _is_sorted(df2[id_col2])
    if use_fast_path:
        print("Using optimized two-pointer comparison (ascending SEC_IDs detected)")
        comparable_cols = intersect_columns(df1, df2, exclude={id_col1, id_col2})

        ids1 = df1[id_col1].to_numpy()
        ids2 = df2[id_col2].to_numpy()
        only_in_1: Set[str] = set()
        only_in_2: Set[str] = set()
        diffs1: Dict[str, List[str]] = {}
        diffs2: Dict[str, List[str]] = {}

        # Build 2D numpy views for comparable columns for vectorized row comparison
        arr1 = df1[comparable_cols].to_numpy(dtype=str, copy=False)
        arr2 = df2[comparable_cols].to_numpy(dtype=str, copy=False)
        col_names = np.array(comparable_cols, dtype=object)

        i = 0
        j = 0
        len1 = len(ids1)
        len2 = len(ids2)
        while i < len1 and j < len2:
            id1 = ids1[i]
            id2 = ids2[j]
            if id1 == id2:
                # Vectorized row comparison
                row_diff_mask = arr1[i] != arr2[j]
                if row_diff_mask.any():
                    diff_cols = col_names[row_diff_mask].tolist()
                    diffs1[id1] = diff_cols
                    diffs2[id2] = diff_cols
                i += 1
                j += 1
            elif id1 < id2:
                only_in_1.add(str(id1))
                i += 1
            else:
                only_in_2.add(str(id2))
                j += 1

        while i < len1:
            only_in_1.add(str(ids1[i]))
            i += 1
        while j < len2:
            only_in_2.add(str(ids2[j]))
            j += 1

        common_ids_count = (len(df1) - len(only_in_1)) if len(df1) <= len(df2) else (len(df2) - len(only_in_2))
        print(
            "Comparison summary: "
            f"common SEC_IDs={common_ids_count}, differing rows={len(diffs1)}, "
            f"only in file1={len(only_in_1)}, only in file2={len(only_in_2)}"
        )

        print(f"Compare elapsed: {time.perf_counter() - start_ts:.2f}s")
        return diffs1, diffs2, only_in_1, only_in_2

    # Fallback: general set/dict-based comparison
    id_map1 = build_row_maps(df1, id_col1)
    id_map2 = build_row_maps(df2, id_col2)

    ids1 = set(id_map1.keys())
    ids2 = set(id_map2.keys())

    common_ids = ids1 & ids2
    only_in_1 = ids1 - ids2
    only_in_2 = ids2 - ids1

    comparable_cols = intersect_columns(df1, df2, exclude={id_col1, id_col2})

    diffs1: Dict[str, List[str]] = {}
    diffs2: Dict[str, List[str]] = {}

    # Vectorize within each row to reduce Python-level loops
    for sec_id in common_ids:
        r1 = df1.loc[id_map1[sec_id], comparable_cols].to_numpy(dtype=str, copy=False)
        r2 = df2.loc[id_map2[sec_id], comparable_cols].to_numpy(dtype=str, copy=False)
        mask = r1 != r2
        if mask.any():
            diff_cols = [col for col, m in zip(comparable_cols, mask) if m]
            diffs1[sec_id] = diff_cols
            diffs2[sec_id] = diff_cols

    print(
        "Comparison summary: "
        f"common SEC_IDs={len(common_ids)}, differing rows={len(diffs1)}, "
        f"only in file1={len(only_in_1)}, only in file2={len(only_in_2)}"
    )
    print(f"Compare elapsed: {time.perf_counter() - start_ts:.2f}s")

    return diffs1, diffs2, only_in_1, only_in_2


def apply_highlights(
    src_path: str,
    out_path: str,
    sec_id_to_diff_cols: Dict[str, List[str]],
    sec_ids_only_in_this: Set[str],
    id_col_name: str,
):
    print(f"Applying highlights for: {src_path}")
    wb = load_workbook(src_path)
    ws = wb.active

    # Map headers to column index (1-based for openpyxl)
    header_row = 1
    headers = {}
    for col_idx, cell in enumerate(ws[header_row], start=1):
        headers[str(cell.value).strip() if cell.value is not None else ""] = col_idx

    # Build SEC_ID -> worksheet row number
    id_col_idx = headers.get(id_col_name)
    if id_col_idx is None:
        # Fallback to first column
        id_col_idx = 1
    sec_id_to_row: Dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row=r, column=id_col_idx).value
        key = str(val).strip() if val is not None else ""
        if key and key not in sec_id_to_row:
            sec_id_to_row[key] = r

    # Convenience header name to column index for diff columns
    diff_col_to_idx: Dict[str, int] = {}
    for col_name in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
        for idx, name in enumerate(col_name, start=1):
            if name is not None:
                diff_col_to_idx[str(name).strip()] = idx

    # 1) Highlight rows that have differences: whole row YELLOW, then specific cells RED
    for sec_id, diff_cols in sec_id_to_diff_cols.items():
        row_num = sec_id_to_row.get(sec_id)
        if not row_num:
            continue
        for c in range(1, ws.max_column + 1):
            ws.cell(row=row_num, column=c).fill = YELLOW_FILL
        for col_name in diff_cols:
            cidx = diff_col_to_idx.get(col_name)
            if cidx is not None:
                ws.cell(row=row_num, column=cidx).fill = RED_FILL

    # 2) Highlight rows only in this workbook: whole row BLUE
    for sec_id in sec_ids_only_in_this:
        row_num = sec_id_to_row.get(sec_id)
        if not row_num:
            continue
        for c in range(1, ws.max_column + 1):
            ws.cell(row=row_num, column=c).fill = BLUE_FILL

    # Ensure output directory exists
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    print(
        f" - Yellow rows: {len(sec_id_to_diff_cols)} | Blue rows: {len(sec_ids_only_in_this)}"
    )
    print(f"Saving highlighted workbook to: {out_path}")
    wb.save(out_path)


def derive_output_paths(file1: str, file2: str, output_dir: str) -> Tuple[str, str]:
    base1 = os.path.splitext(os.path.basename(file1))[0]
    base2 = os.path.splitext(os.path.basename(file2))[0]
    out1 = os.path.join(output_dir, f"{base1}__highlighted.xlsx")
    out2 = os.path.join(output_dir, f"{base2}__highlighted.xlsx")
    return out1, out2


def _auto_discover_input_files(input_dir: str) -> Tuple[str, str]:
    candidates: List[str] = []
    if os.path.isdir(input_dir):
        for name in sorted(os.listdir(input_dir)):
            if name.lower().endswith(".xlsx") and not name.startswith("~$"):
                candidates.append(os.path.join(input_dir, name))
    if len(candidates) < 2:
        raise FileNotFoundError(
            f"Expected at least two .xlsx files in '{input_dir}'. Found: {len(candidates)}"
        )
    return candidates[0], candidates[1]


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Compare two Excel files by SEC_ID (first column or column named SEC_ID). "
            "Highlight differing cells red and full differing rows yellow. Rows with a SEC_ID only present in one file are highlighted blue in that file."
        )
    )
    parser.add_argument(
        "file1",
        nargs="?",
        help="Path to first Excel file (.xlsx). If omitted, auto-uses two files from ./input",
    )
    parser.add_argument(
        "file2",
        nargs="?",
        help="Path to second Excel file (.xlsx). If omitted, auto-uses two files from ./input",
    )
    parser.add_argument(
        "--input",
        "-i",
        default=os.path.join("input"),
        help="Directory to read input Excel files when auto-discovering (default: input)",
    )
    parser.add_argument(
        "--output",
        "-o",
        default=os.path.join("output"),
        help="Directory to write highlighted Excel files (default: output)",
    )

    args = parser.parse_args()

    if args.file1 and args.file2:
        file1 = args.file1
        file2 = args.file2
    else:
        try:
            file1, file2 = _auto_discover_input_files(args.input)
            print(f"Auto-discovered input files: {file1} | {file2}")
        except Exception as e:
            print(f"Input discovery error: {e}")
            return 1

    if not os.path.exists(file1):
        print(f"Error: file not found: {file1}")
        return 1
    if not os.path.exists(file2):
        print(f"Error: file not found: {file2}")
        return 1

    try:
        df1 = read_excel_as_str(file1)
        df2 = read_excel_as_str(file2)
    except Exception as e:
        print(f"Failed to read Excel files: {e}")
        return 1

    id_col1 = get_sec_id_column_name(df1)
    id_col2 = get_sec_id_column_name(df2)
    print(f"Detected SEC_ID columns -> file1: '{id_col1}', file2: '{id_col2}'")

    diffs1, diffs2, only_in_1, only_in_2 = compare_rows(df1, df2, id_col1, id_col2)

    out1, out2 = derive_output_paths(file1, file2, args.output)

    try:
        apply_highlights(file1, out1, diffs1, only_in_1, id_col1)
        apply_highlights(file2, out2, diffs2, only_in_2, id_col2)
    except Exception as e:
        print(f"Failed to write highlighted workbooks: {e}")
        return 1

    print(f"Done. Wrote: {out1}")
    print(f"Done. Wrote: {out2}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
